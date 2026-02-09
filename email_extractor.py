import win32com.client
from pathlib import Path
from datetime import datetime
import os

def search_outlook_emails(sender_email: str, search_term: str, download_folder: str, workbook_dir: str = None) -> list[str]:
    """Search Outlook for emails from sender with search_term in subject.
    
    Args:
        sender_email: Email address to search for
        search_term: Term to search in subject line
        download_folder: Relative folder path to save PDF attachments
        workbook_dir: Directory of workbook for absolute path resolution
        
    Returns:
        List of relative paths to downloaded PDF files
    """
    try:
        # Connect to Outlook
        outlook = win32com.client.Dispatch("Outlook.Application")
        namespace = outlook.GetNamespace("MAPI")
        inbox = namespace.GetDefaultFolder(6)  # 6 = Inbox
        
        print(f"  Searching Outlook for emails from: {sender_email}")
        print(f"  Looking for subject containing: {search_term}")
        
        # Get all items in inbox (we'll filter manually for better compatibility)
        items = inbox.Items
        items.Sort("[ReceivedTime]", True)  # Sort by most recent first
        
        downloaded_files = []
        emails_checked = 0
        emails_matched = 0
        
        # Loop through all emails and filter manually
        for mail in items:
            try:
                emails_checked += 1
                
                # Only process MailItem objects
                if mail.Class != 43:  # 43 = olMail
                    continue
                
                # Get subject
                subject = getattr(mail, 'Subject', '')
                if not subject:
                    continue
                
                # Check if search term is in subject
                if search_term.lower() not in subject.lower():
                    continue
                
                # Get sender email - try multiple properties
                sender = None
                try:
                    # Try SenderEmailAddress first
                    sender = getattr(mail, 'SenderEmailAddress', '')
                    
                    # If it's Exchange format (starts with /), try to get SMTP address
                    if sender and sender.startswith('/'):
                        try:
                            sender_obj = mail.Sender
                            if sender_obj:
                                sender = sender_obj.GetExchangeUser().PrimarySmtpAddress
                        except:
                            pass
                    
                    # Also try SenderName property
                    if not sender or '@' not in sender:
                        sender = getattr(mail, 'SenderName', '')
                except:
                    sender = ''
                
                print(f"  Checking: {subject[:40]}... (from: {sender})")
                
                # Check if sender matches (case-insensitive)
                if sender_email.lower() not in sender.lower():
                    continue
                
                emails_matched += 1
                print(f"  [MATCH] Found email: {subject[:50]}...")
                
                # Process attachments
                attachments = mail.Attachments
                if attachments.Count == 0:
                    print(f"    No attachments found")
                    continue
                
                for i in range(1, attachments.Count + 1):
                    attachment = attachments.Item(i)
                    filename = attachment.FileName
                    
                    if filename.lower().endswith('.pdf'):
                        # Create unique filename
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
                        save_filename = f"{timestamp}_{filename}"
                        
                        # Use absolute path for Outlook SaveAsFile
                        if workbook_dir:
                            abs_save_path = os.path.abspath(os.path.join(workbook_dir, download_folder, save_filename))
                        else:
                            abs_save_path = os.path.abspath(os.path.join(download_folder, save_filename))
                        
                        # Save attachment
                        attachment.SaveAsFile(abs_save_path)
                        
                        # Store relative path
                        relative_path = os.path.join(download_folder, save_filename)
                        downloaded_files.append(relative_path)
                        print(f"    [OK] Downloaded: {filename}")
                    else:
                        print(f"    Skipped (not PDF): {filename}")
                
                # Limit search to recent emails
                if emails_checked >= 100:
                    break
                    
            except Exception as e:
                print(f"    Error processing email: {e}")
                continue
        
        print(f"  Checked {emails_checked} emails, matched {emails_matched}")
        return downloaded_files
        
    except Exception as e:
        print(f"  ERROR accessing Outlook: {e}")
        import traceback
        traceback.print_exc()
        return []

def get_tips_info(workbook, fh_number: str) -> dict:
    """Lookup FH number in TIPS sheet and return row data.
    
    Args:
        workbook: Openpyxl workbook object
        fh_number: FH number to lookup
        
    Returns:
        Dictionary with MY_TIPS, PASSWORD1, PASSWORD2, PASSWORD3, FUND_HOUSE
    """
    if "TIPS" not in workbook.sheetnames:
        return {}
    
    ws = workbook["TIPS"]
    
    # Search for FH number in column 1 (Nº)
    for row_num in range(2, ws.max_row + 1):
        cell_value = ws.cell(row_num, 1).value
        if cell_value and str(cell_value).strip() == str(fh_number).strip():
            # Found the row
            return {
                "MY_TIPS": ws.cell(row_num, 17).value or "",  # Column 17
                "PASSWORD1": ws.cell(row_num, 18).value or "",  # Column 18
                "PASSWORD2": ws.cell(row_num, 19).value or "",  # Column 19
                "PASSWORD3": ws.cell(row_num, 20).value or "",  # Column 20
                "FUND_HOUSE": ws.cell(row_num, 2).value or "",  # Column 2
            }
    
    return {}

def process_opc_rows(workbook_path: str) -> int:
    """Process OPC rows: search emails, download PDFs, log to CN Database.
    
    Args:
        workbook_path: Path to the Excel workbook
        
    Returns:
        Number of PDFs downloaded
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(workbook_path, keep_vba=True)
    
    # Check required sheets exist
    required_sheets = ["OPC", "TIPS", "Parameters", "CN Database"]
    for sheet in required_sheets:
        if sheet not in wb.sheetnames:
            print(f"ERROR: Required sheet '{sheet}' not found")
            return 0
    
    ws_opc = wb["OPC"]
    ws_params = wb["Parameters"]
    ws_cn = wb["CN Database"]
    
    # Get configuration
    confirmation_email = ws_params["D2"].value
    if not confirmation_email:
        print("ERROR: Confirmation email not found in Parameters D2")
        return 0
    
    # Get download folder from Parameters or use default
    download_folder = None
    for row_num in range(1, ws_params.max_row + 1):
        if ws_params.cell(row_num, 1).value == "DOWNLOAD_FOLDER":
            download_folder = ws_params.cell(row_num, 2).value
            break
    
    if not download_folder:
        download_folder = "downloads"
    
    # Get workbook directory
    workbook_dir = Path(workbook_path).parent
    
    # Create downloads folder if needed (relative to workbook location)
    download_path = workbook_dir / download_folder
    download_path.mkdir(parents=True, exist_ok=True)
    
    print(f"Confirmation email: {confirmation_email}")
    print(f"Download folder: {download_folder}")
    print("\n" + "="*60)
    
    total_pdfs = 0
    
    # Process each OPC row
    for row_num in range(2, ws_opc.max_row + 1):
        fh_value = ws_opc.cell(row_num, 5).value  # Column 5 = FH
        
        if not fh_value:
            continue
        
        fh_value = str(fh_value).strip()
        print(f"\nProcessing OPC row {row_num}: FH = {fh_value}")
        
        # Get TIPS info for this FH
        tips_info = get_tips_info(wb, fh_value)
        
        if not tips_info or not tips_info.get("MY_TIPS"):
            print(f"  WARNING: No TIPS info found for FH {fh_value}")
            continue
        
        my_tips = tips_info["MY_TIPS"].strip().strip('"')
        print(f"  MY TIPS search term: {my_tips}")
        
        # Search Outlook and download PDFs
        downloaded_files = search_outlook_emails(
            confirmation_email,
            my_tips,
            download_folder,
            str(workbook_dir)
        )
        
        # Log each PDF to CN Database
        for pdf_path in downloaded_files:
            # Find next empty row in CN Database
            cn_row = ws_cn.max_row + 1
            
            # Write to CN Database
            ws_cn.cell(cn_row, 1).value = cn_row - 1  # ID
            ws_cn.cell(cn_row, 2).value = pdf_path  # File Path
            
            # Store passwords in temp columns for Python processing
            ws_cn.cell(cn_row, 16).value = tips_info.get("PASSWORD1", "")
            ws_cn.cell(cn_row, 17).value = tips_info.get("PASSWORD2", "")
            ws_cn.cell(cn_row, 18).value = tips_info.get("PASSWORD3", "")
            
            total_pdfs += 1
            print(f"  Logged to CN Database row {cn_row}")
    
    # Save workbook
    wb.save(workbook_path)
    print("\n" + "="*60)
    print(f"[SUCCESS] Downloaded and logged {total_pdfs} PDFs to CN Database")
    
    return total_pdfs
