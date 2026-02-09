import sys
from pathlib import Path
from datetime import datetime

from config import EXTRACTION_PROMPT, LLM_MODEL
from pdf_processor import try_open_pdf, rasterize_pages
from ocr_reader import extract_text_from_images
from data_extractor import extract_structured_data
from email_extractor import process_opc_rows

def process_pdfs_from_cn_database(workbook_path: str) -> int:
    """Process PDFs based on CN Database rows (OPC-driven workflow).
    
    For each row in CN Database:
    1. Read File Path (col 2) and passwords (temp cols 16-18)
    2. Open PDF with row-specific passwords
    3. Extract text and send to LLM
    4. Update CN Database row with extracted data
    
    Returns: Number of successfully processed PDFs
    """
    if not workbook_path or not Path(workbook_path).exists():
        print("Error: Workbook path not provided or doesn't exist")
        return 0
    
    try:
        from openpyxl import load_workbook
        wb = load_workbook(workbook_path, keep_vba=True)
        
        if "CN Database" not in wb.sheetnames:
            print("Error: CN Database sheet not found")
            return 0
        
        ws = wb["CN Database"]
        processed_count = 0
        workbook_dir = Path(workbook_path).parent
        
        # Process each row (starting from row 2, assuming row 1 is headers)
        for row_num in range(2, ws.max_row + 1):
            file_path = ws.cell(row_num, 2).value  # Column 2 = File Path
            
            if not file_path:
                continue
            
            # Convert relative path to absolute based on workbook location
            if not Path(file_path).is_absolute():
                file_path = str(workbook_dir / file_path)
            
            if not Path(file_path).exists():
                print(f"Row {row_num}: File not found: {file_path}")
                continue
            
            # Get row-specific passwords from temp columns
            pwd1 = ws.cell(row_num, 16).value or ""  # Temp col 16
            pwd2 = ws.cell(row_num, 17).value or ""  # Temp col 17
            pwd3 = ws.cell(row_num, 18).value or ""  # Temp col 18
            passwords = [p for p in [pwd1, pwd2, pwd3] if p]
            
            print(f"\nProcessing row {row_num}: {Path(file_path).name}")
            
            try:
                # Open PDF with row-specific passwords
                doc = try_open_pdf(file_path, passwords)
                if doc is None:
                    print(f"  ERROR: Could not decrypt PDF")
                    ws.cell(row_num, 3).value = "ERROR: Password protected"
                    continue
                
                # Extract text (rasterize + OCR)
                print(f"  Rasterizing {len(doc)} pages...")
                images = rasterize_pages(doc)
                doc.close()
                
                print(f"  Running OCR...")
                text = extract_text_from_images(images)
                
                # Extract CN data with LLM
                print(f"  Extracting CN data with LLM...")
                data = extract_structured_data(text, EXTRACTION_PROMPT, LLM_MODEL)
                
                # Update CN Database row with extracted data
                ws.cell(row_num, 3).value = data.get("is_cn", "")
                ws.cell(row_num, 4).value = data.get("operation_type", "")
                ws.cell(row_num, 5).value = data.get("is_multiseries", "")
                ws.cell(row_num, 6).value = data.get("currency", "")
                ws.cell(row_num, 7).value = data.get("gross_amount", "")
                ws.cell(row_num, 8).value = data.get("net_amount", "")
                ws.cell(row_num, 9).value = data.get("units", "")
                ws.cell(row_num, 10).value = data.get("equalization", "")
                ws.cell(row_num, 11).value = data.get("fees", "")
                ws.cell(row_num, 12).value = data.get("nav_price", "")
                ws.cell(row_num, 13).value = data.get("nav_date", "")
                ws.cell(row_num, 14).value = data.get("settlement_date", "")
                
                # Clear temp password columns
                ws.cell(row_num, 16).value = None
                ws.cell(row_num, 17).value = None
                ws.cell(row_num, 18).value = None
                
                processed_count += 1
                print(f"  [OK] Successfully extracted CN data")
                
            except Exception as e:
                print(f"  ERROR: {e}")
                ws.cell(row_num, 3).value = f"ERROR: {str(e)}"
        
        # Save workbook
        wb.save(workbook_path)
        print(f"\n[OK] Workbook saved. Processed {processed_count} PDFs successfully.")
        return processed_count
        
    except Exception as e:
        print(f"Fatal error processing CN Database: {e}")
        return 0

def main():
    """Main entry point for OPC-driven CN extraction workflow.
    
    Workflow:
    1. Python reads OPC sheet rows
    2. For each FH value, Python searches Outlook and downloads PDFs
    3. Python logs File Path and passwords to CN Database
    4. Python processes each CN Database row with row-specific passwords
    5. Python updates CN Database with extracted CN data
    """
    # Get workbook path from command line
    workbook_path = None
    if len(sys.argv) > 1:
        workbook_path = sys.argv[1]
    else:
        workbook_path = Path(__file__).parent / "OPC_TEST.xlsm"
    
    if not workbook_path or not Path(workbook_path).exists():
        print(f"Error: OPC_TEST workbook not found: {workbook_path}")
        print("Usage: python main.py <path_to_OPC_TEST.xlsm>")
        return 1
    
    print("="*60)
    print("OPC-DRIVEN CN EXTRACTION WORKFLOW")
    print("="*60)
    print(f"Workbook: {workbook_path}")
    print("\nSTEP 1: Email Extraction and PDF Download")
    print("="*60)
    
    # Step 1: Process OPC rows and download PDFs from Outlook
    try:
        pdf_count = process_opc_rows(workbook_path)
        
        if pdf_count == 0:
            print("\n[WARNING] No PDFs were downloaded from emails.")
            print("Check that:")
            print("  - Outlook is running")
            print("  - Emails exist with the MY TIPS search terms")
            print("  - Confirmation email address is correct in Parameters D2")
            return 1
    except Exception as e:
        print(f"\n[ERROR] Failed to process emails: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    print(f"\n\nSTEP 2: PDF Processing (OCR + LLM Extraction)")
    print("="*60)
    
    # Step 2: Process PDFs with OCR and LLM
    processed_count = process_pdfs_from_cn_database(workbook_path)
    
    if processed_count > 0:
        print("\n" + "="*60)
        print(f"[SUCCESS] Complete! Processed {processed_count} PDFs!")
        print("="*60)
        print("CN Database has been updated with extracted data.")
        return 0
    else:
        print("\n[WARNING] No PDFs were processed successfully.")
        return 1

if __name__ == "__main__":
    sys.exit(main())
