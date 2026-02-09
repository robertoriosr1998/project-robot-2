from openpyxl import load_workbook

wb = load_workbook('OPC_TEST.xlsm', keep_vba=True)
ws = wb['CN Database']

print('CN Database contents:')
print('='*80)
for row_num in range(1, min(5, ws.max_row + 1)):
    print(f'\nRow {row_num}:')
    print(f'  ID: {ws.cell(row_num, 1).value}')
    print(f'  File Path: {ws.cell(row_num, 2).value}')
    print(f'  Is it a CN?: {ws.cell(row_num, 3).value}')
    print(f'  Operation Type: {ws.cell(row_num, 4).value}')
    print(f'  Currency: {ws.cell(row_num, 6).value}')
    print(f'  Gross Amount: {ws.cell(row_num, 7).value}')
    print(f'  Net Amount: {ws.cell(row_num, 8).value}')
